# invoice_generator/generate_invoice.py
# Main script to orchestrate invoice generation using a processor-based pattern.

import os
import json
import pickle
import argparse
import shutil
import openpyxl
import traceback
import sys
import time
import datetime
import logging
from pathlib import Path
from typing import Optional, Dict, Any, List
import ast
import re

# --- Import utility functions from the new structure ---
from .config.config_loader import BundledConfigLoader
from .builders.text_replacement_builder import TextReplacementBuilder
from .builders.workbook_builder import WorkbookBuilder
from .processors.single_table_processor import SingleTableProcessor
from .processors.multi_table_processor import MultiTableProcessor

logger = logging.getLogger(__name__)

# --- Helper Functions (derive_paths, load_config, load_data) ---
# These functions remain largely the same but are part of the new script.
def derive_paths(input_data_path_str: str, template_dir_str: str, config_dir_str: str) -> Optional[Dict[str, Path]]:
    """
    Derives template and config file paths based on the input data filename.
    """
    logger.info(f"Deriving paths from input: {input_data_path_str}")
    try:
        input_data_path = Path(input_data_path_str).resolve()
        template_dir = Path(template_dir_str).resolve()
        config_dir = Path(config_dir_str).resolve()

        if not input_data_path.is_file(): 
            logger.error(f"Input data file not found: {input_data_path}")
            return None
        if not template_dir.is_dir(): 
            logger.error(f"Template directory not found: {template_dir}")
            return None
        if not config_dir.is_dir(): 
            logger.error(f"Config directory not found: {config_dir}")
            return None

        base_name = input_data_path.stem
        template_name_part = base_name
        suffixes_to_remove = ['_data', '_input', '_pkl']
        prefixes_to_remove = ['data_']

        for suffix in suffixes_to_remove:
            if base_name.lower().endswith(suffix):
                template_name_part = base_name[:-len(suffix)]
                break
        else:
            for prefix in prefixes_to_remove:
                if base_name.lower().startswith(prefix):
                    template_name_part = base_name[len(prefix):]
                    break

        if not template_name_part:
            logger.error(f"Could not derive template name part from: '{base_name}'")
            return None
        logger.debug(f"Derived initial template name part: '{template_name_part}'")

        exact_template_filename = f"{template_name_part}.xlsx"
        exact_config_filename = f"{template_name_part}_config.json"
        exact_template_path = template_dir / exact_template_filename
        exact_config_path = config_dir / exact_config_filename
        logger.debug(f"Checking for exact match: Template='{exact_template_path}', Config='{exact_config_path}'")

        if exact_template_path.is_file() and exact_config_path.is_file():
            logger.info("Found exact match for template and config")
            return {"data": input_data_path, "template": exact_template_path, "config": exact_config_path}
        
        # Check for bundled config subdirectory pattern: config_bundled/JF_config/JF_config.json
        bundled_config_subdir = config_dir / exact_config_filename.replace('.json', '') / exact_config_filename
        if exact_template_path.is_file() and bundled_config_subdir.is_file():
            logger.info(f"Found exact match for template and bundled config (subdir): {bundled_config_subdir}")
            return {"data": input_data_path, "template": exact_template_path, "config": bundled_config_subdir}
        
        logger.debug("Exact match not found. Attempting prefix matching...")
        prefix_match = re.match(r'^([a-zA-Z]+[-_]?[a-zA-Z]*)', template_name_part)
        if prefix_match:
            logger.debug("Exact match not found. Attempting prefix matching...")
            prefix_match = re.match(r'^([a-zA-Z]+[-_]?[a-zA-Z]*)', template_name_part)
            if prefix_match:
                prefix = prefix_match.group(1)
                logger.debug(f"Extracted prefix: '{prefix}'")
                prefix_template_filename = f"{prefix}.xlsx"
                prefix_config_filename = f"{prefix}_config.json"
                prefix_template_path = template_dir / prefix_template_filename
                prefix_config_path = config_dir / prefix_config_filename
                logger.debug(f"Checking for prefix match: Template='{prefix_template_path}', Config='{prefix_config_path}'")

                if prefix_template_path.is_file() and prefix_config_path.is_file():
                    logger.info("Found prefix match for template and config")
                    return {"data": input_data_path, "template": prefix_template_path, "config": prefix_config_path}
                else:
                    logger.debug("Prefix match not found")
            else:
                logger.debug("Could not extract a letter-based prefix")

            logger.error(f"Could not find matching template/config files using exact ('{template_name_part}') or prefix methods")
            if not exact_template_path.is_file(): 
                logger.error(f"Template file not found: {exact_template_path}")
            if not exact_config_path.is_file(): 
                logger.error(f"Configuration file not found: {exact_config_path}")
            return None

    except Exception as e:
        logger.error(f"Error deriving file paths: {e}")
        traceback.print_exc()
        return None

def load_config(config_path: Path) -> Optional[Dict[str, Any]]:
    """Loads and parses the JSON configuration file."""
    logger.info(f"Loading configuration from: {config_path}")
    try:
        with open(config_path, 'r', encoding='utf-8') as f: config_data = json.load(f)
        logger.info("Configuration loaded successfully.")
        # Detect bundled config version
        if '_meta' in config_data and 'config_version' in config_data['_meta']:
            logger.info(f"Detected bundled config version: {config_data['_meta']['config_version']}")
        return config_data
    except Exception as e:
        logger.error(f"Error loading configuration file {config_path}: {e}")
        traceback.print_exc()
        return None

def build_sheet_config_from_bundled(config: Dict[str, Any], sheet_name: str) -> Dict[str, Any]:
    """
    Builds a sheet configuration dictionary from bundled config format.
    
    Bundled format has config split across:
    - processing.data_sources[sheet_name]: data source type
    - styling_bundle[sheet_name]: styling config
    - layout_bundle[sheet_name]: layout, structure, data_flow, content, footer
    - defaults: global defaults to merge with sheet-specific config
    
    Returns a unified sheet_config dict that processors can use.
    """
    sheet_config = {}
    
    # Get data source type
    if 'processing' in config and 'data_sources' in config['processing']:
        sheet_config['data_source_type'] = config['processing']['data_sources'].get(sheet_name)
    
    # Get styling config
    if 'styling_bundle' in config and sheet_name in config['styling_bundle']:
        sheet_config['styling_config'] = config['styling_bundle'][sheet_name]
    
    # Get layout config (structure, data_flow, content, footer)
    if 'layout_bundle' in config and sheet_name in config['layout_bundle']:
        layout = config['layout_bundle'][sheet_name]
        sheet_config['sheet_config'] = layout.get('structure', {})
        sheet_config['mappings'] = layout.get('data_flow', {}).get('mappings', {})
        sheet_config['static_content'] = layout.get('content', {})
        sheet_config['footer_config'] = layout.get('footer', {})
    
    # Merge global defaults
    if 'defaults' in config:
        defaults = config['defaults']
        # Merge footer defaults if not specified
        if 'footer_config' in sheet_config and 'footer' in defaults:
            for key, value in defaults['footer'].items():
                sheet_config['footer_config'].setdefault(key, value)
    
    return sheet_config

def load_data(data_path: Path) -> Optional[Dict[str, Any]]:
    """ Loads and parses the input data file. Supports .json and .pkl. """
    logger.info(f"Loading data from: {data_path}")
    invoice_data = None; file_suffix = data_path.suffix.lower()
    try:
        if file_suffix == '.json':
            with open(data_path, 'r', encoding='utf-8') as f: invoice_data = json.load(f)
        elif file_suffix == '.pkl':
            with open(data_path, 'rb') as f: invoice_data = pickle.load(f)
        else:
            logger.error(f"Error: Unsupported data file extension: '{file_suffix}'.")
        
        # Key conversion logic remains the same
        for key_to_convert in ["standard_aggregation_results", "custom_aggregation_results"]:
            raw_data = invoice_data.get(key_to_convert)
            if isinstance(raw_data, dict):
                processed_data = {}
                decimal_pattern = re.compile(r"Decimal\('(-?\d*\.?\d+)'\)")
                for key_str, value_dict in raw_data.items():
                    try:
                        processed_key_str = decimal_pattern.sub(r"'\1'", key_str)
                        key_tuple = ast.literal_eval(processed_key_str)
                        processed_data[key_tuple] = value_dict
                    except Exception as e:
                        logger.warning(f"Warning: Could not convert key '{key_str}': {e}")
                invoice_data[key_to_convert] = processed_data
        
        return invoice_data
    except Exception as e:
        logger.error(f"Error loading data file {data_path}: {e}")
        traceback.print_exc()
        return None

def generate_metadata(output_path: Path, status: str, execution_time: float, sheets_processed: List[str], sheets_failed: List[str], error_message: Optional[str] = None, invoice_data: Optional[Dict] = None, cli_args: Optional[argparse.Namespace] = None, replacements_log: Optional[List[Dict]] = None, header_info: Optional[Dict] = None):
    """Generates a metadata JSON file for backend integration."""
    
    # 1. Basic Info
    metadata = {
        "status": status,
        "output_file": str(output_path),
        "timestamp": datetime.datetime.now().isoformat(),
        "execution_time": execution_time,
        "sheets_processed": sheets_processed,
        "sheets_failed": sheets_failed,
        "error_message": error_message
    }

    # 2. Config Info
    if cli_args:
        metadata["config_info"] = {
            "daf_mode": cli_args.DAF,
            "custom_mode": cli_args.custom,
            "input_file": Path(cli_args.input_data_file).name,
            "config_dir": cli_args.configdir
        }

    # 3. Database Export (Packing List Items)
    if invoice_data and "processed_tables_data" in invoice_data:
        processed_data = invoice_data["processed_tables_data"]
        packing_list_items = []
        
        # Iterate through tables (e.g., "1", "2")
        for table_id, table_data in processed_data.items():
            # All lists in table_data should be same length
            row_count = len(table_data.get("po", []))
            
            for i in range(row_count):
                try:
                    item = {
                        "po": table_data.get("po", [])[i],
                        "item": table_data.get("item", [])[i],
                        "description": table_data.get("description", [])[i],
                        "pcs": table_data.get("pcs", [])[i],
                        "sqft": table_data.get("sqft", [])[i],
                        "pallet_count": table_data.get("pallet_count", [])[i],
                        "net": table_data.get("net", [])[i],
                        "gross": table_data.get("gross", [])[i],
                        "cbm": table_data.get("cbm", [])[i]
                    }
                    packing_list_items.append(item)
                except IndexError:
                    continue # Skip malformed rows

        # 4. Summary Statistics
        total_pcs = sum(int(i["pcs"]) for i in packing_list_items if str(i["pcs"]).isdigit())
        total_sqft = sum(float(i["sqft"]) for i in packing_list_items if str(i["sqft"]).replace('.', '', 1).isdigit())
        total_pallets = sum(int(i["pallet_count"]) for i in packing_list_items if str(i["pallet_count"]).isdigit())
        
        # Calculate total amount (need amount column which might be missing in item dict if not added above)
        # Let's add amount to item dict above first? Or just calculate from invoice_data if available.
        # Better to stick to what we extracted.
        
        metadata["database_export"] = {
            "summary": {
                "total_pcs": total_pcs,
                "total_sqft": total_sqft,
                "total_pallets": total_pallets,
                "item_count": len(packing_list_items)
            },
            "packing_list_items": packing_list_items
        }

    # 5. Invoice Terms (from text replacements)
    if replacements_log:
        terms_found = set()
        for entry in replacements_log:
            if entry.get("term"):
                terms_found.add(entry["term"])
        
        metadata["invoice_terms"] = {
            "detected_terms": list(terms_found),
            "replacements_detail": replacements_log
        }

    # 6. Header Info (Company Name, Address)
    if header_info:
        metadata["header_info"] = header_info

    # 7. Invoice Info (No, Date, Ref)
    if invoice_data and "processed_tables_data" in invoice_data:
        try:
            # Assuming data is in the first table "1"
            table_data = invoice_data["processed_tables_data"].get("1", {})
            
            # Helper to safely get the first value from a list
            def get_first(key):
                val = table_data.get(key)
                if isinstance(val, list) and val:
                    return val[0]
                return None

            metadata["invoice_info"] = {
                "inv_no": get_first("inv_no"),
                "inv_date": get_first("inv_date"),
                "inv_ref": get_first("inv_ref")
            }
        except Exception as e:
            # Don't fail metadata generation if extraction fails
            metadata["invoice_info_error"] = str(e)

    meta_path = output_path.parent / (output_path.name + ".meta.json")
    try:
        with open(meta_path, "w", encoding="utf-8") as f:
            json.dump(metadata, f, indent=2)
        logger.info(f"Metadata generated: {meta_path}")
    except Exception as e:
        logger.error(f"Failed to generate metadata: {e}")

def main():
    """Main function to orchestrate invoice generation."""
    start_time = time.time()
    
    parser = argparse.ArgumentParser(description="Generate Invoice from Template and Data using configuration files.")
    parser.add_argument("input_data_file", help="Path to the input data file (.json or .pkl).")
    parser.add_argument("-o", "--output", default="result.xlsx", help="Path for the output Excel file.")
    parser.add_argument("-t", "--templatedir", default="./TEMPLATE", help="Directory containing template Excel files.")
    parser.add_argument("-c", "--configdir", default="./configs", help="Directory containing configuration JSON files.")
    parser.add_argument("--DAF", action="store_true", help="Enable DAF-specific processing.")
    parser.add_argument("--custom", action="store_true", help="Enable custom processing logic.")
    parser.add_argument("--debug", action="store_true", help="Enable debug logging (shows all DEBUG messages).")
    parser.add_argument("--log-level", choices=['DEBUG', 'INFO', 'WARNING', 'ERROR', 'CRITICAL'], 
                        default='INFO', help="Set logging level (default: INFO).")
    args = parser.parse_args()
    
    # Configure logging based on command-line arguments
    if args.debug:
        log_level = logging.DEBUG
    else:
        log_level = getattr(logging, args.log_level, logging.INFO)
    
    # Configure logging with separate handlers and colors
    # INFO/DEBUG -> stdout (white), WARNING -> stderr (yellow), ERROR/CRITICAL -> stderr (red)
    logger = logging.getLogger()
    logger.setLevel(log_level)
    logger.handlers.clear()  # Clear any existing handlers
    
    # Custom formatter with ANSI color codes and level-based format
    class ColoredFormatter(logging.Formatter):
        COLORS = {
            'DEBUG': '\033[36m',      # Cyan
            'INFO': '\033[0m',         # White (reset)
            'WARNING': '\033[33m',     # Yellow
            'ERROR': '\033[31m',       # Red
            'CRITICAL': '\033[35m',    # Magenta
            'RESET': '\033[0m'         # Reset
        }
        
        # Different formats for different levels
        FORMATS = {
            logging.DEBUG: '>>> %(levelname)s >>> [%(filename)s:%(lineno)d in %(funcName)s()] %(message)s',
            logging.INFO: '>>> %(levelname)s >>> %(message)s',  # Clean format for INFO
            logging.WARNING: '>>> %(levelname)s >>> [%(filename)s:%(lineno)d in %(funcName)s()] %(message)s',
            logging.ERROR: '>>> %(levelname)s >>> [%(filename)s:%(lineno)d in %(funcName)s()] %(message)s',
            logging.CRITICAL: '>>> %(levelname)s >>> [%(filename)s:%(lineno)d in %(funcName)s()] %(message)s',
        }
        
        def format(self, record):
            # Apply color to level name
            color = self.COLORS.get(record.levelname, self.COLORS['RESET'])
            record.levelname = f"{color}{record.levelname}{self.COLORS['RESET']}"
            
            # Use format based on log level
            log_fmt = self.FORMATS.get(record.levelno, self.FORMATS[logging.INFO])
            formatter = logging.Formatter(log_fmt)
            return formatter.format(record)
    
    # Create formatter instance
    formatter = ColoredFormatter()
    
    # Handler for INFO and DEBUG -> stdout (normal color)
    stdout_handler = logging.StreamHandler(sys.stdout)
    stdout_handler.setLevel(logging.DEBUG)
    stdout_handler.setFormatter(formatter)
    stdout_handler.addFilter(lambda record: record.levelno <= logging.INFO)
    logger.addHandler(stdout_handler)
    
    # Handler for WARNING, ERROR, CRITICAL -> stderr (colored: yellow/red)
    stderr_handler = logging.StreamHandler(sys.stderr)
    stderr_handler.setLevel(logging.WARNING)
    stderr_handler.setFormatter(formatter)
    logger.addHandler(stderr_handler)
    
    logger = logging.getLogger(__name__)
    logger.info("=== Starting Invoice Generation (Refactored) ===")
    logger.info(f"Started at: {time.strftime('%H:%M:%S', time.localtime(start_time))}")
    logger.debug(f"Arguments: {vars(args)}")

    paths = derive_paths(args.input_data_file, args.templatedir, args.configdir)
    if not paths: sys.exit(1)

    # Use BundledConfigLoader instead of raw JSON loading
    try:
        config_loader = BundledConfigLoader(paths['config'])
    except Exception as e:
        logger.error(f"Failed to load configuration: {e}")
        sys.exit(1)
    
    invoice_data = load_data(paths['data'])
    if not invoice_data: sys.exit(1)

    output_path = Path(args.output).resolve()
    try:
        output_path.parent.mkdir(parents=True, exist_ok=True)
    except Exception as e:
        logger.error(f"Error creating output directory: {e}")
        sys.exit(1)
    
    # NOTE: We no longer copy the template file directly.
    # Instead, we'll load it as read-only and create a new workbook separately.
    logger.info("Template will be loaded as read-only for state capture")

    logger.info("Loading template and creating new workbook...")
    template_workbook = None
    output_workbook = None
    processing_successful = True

    try:
        # Step 1: Load template workbook as read_only=False
        # - read_only=False: Allows access to merged_cells for state capture
        # - data_only=False (default): Preserves formulas and text with '=' signs
        logger.debug(f"Loading template from: {paths['template']}")
        template_workbook = openpyxl.load_workbook(
            paths['template'], 
            read_only=False
        )
        logger.debug(f"Template loaded successfully (read_only=False)")
        logger.debug(f"   Template sheets: {template_workbook.sheetnames}")
        
        # Step 2: Collect all sheet names from template
        template_sheet_names = template_workbook.sheetnames
        logger.debug(f"Found {len(template_sheet_names)} sheets in template")
        
        # Step 3: Create WorkbookBuilder with template sheet names
        logger.debug("Creating new output workbook using WorkbookBuilder...")
        workbook_builder = WorkbookBuilder(sheet_names=template_sheet_names)
        
        # Step 4: Build the new clean workbook
        output_workbook = workbook_builder.build()
        logger.debug(f"New output workbook created with {len(output_workbook.sheetnames)} sheets")
        
        # Step 5: Store references to both workbooks
        # - template_workbook: Used for reading template state (READ-ONLY usage)
        # - output_workbook: Used for writing final output (WRITABLE)
        workbook = output_workbook  # Keep 'workbook' name for compatibility with rest of code
        logger.debug("Both template (read) and output (write) workbooks ready")

        # Get sheets to process from config loader
        sheets_to_process_config = config_loader.get_sheets_to_process()
        sheets_to_process = [s for s in sheets_to_process_config if s in workbook.sheetnames]

        if not sheets_to_process:
            logger.error("Error: No valid sheets to process."); sys.exit(1)

        # Text replacement now happens at template state level (in LayoutBuilder/processors)
        # This is more efficient and applies to the captured state before restoration
        # Old workbook-level replacement is disabled to avoid double-replacement

        # Global pallet calculation remains the same
        final_grand_total_pallets = 0
        processed_tables_data_for_calc = invoice_data.get('processed_tables_data', {})
        if isinstance(processed_tables_data_for_calc, dict):
            # Simplified calculation
            final_grand_total_pallets = sum(int(c) for t in processed_tables_data_for_calc.values() for c in t.get("pallet_count", []) if str(c).isdigit())
        logger.debug(f"DEBUG: Globally calculated final grand total pallets: {final_grand_total_pallets}")

        # --- REFACTORED Main Processing Loop ---
        sheets_processed = []
        sheets_failed = []
        
        for sheet_name in sheets_to_process:
            logger.info(f"Processing sheet '{sheet_name}'")
            if sheet_name not in workbook.sheetnames:
                logger.warning(f"Warning: Sheet '{sheet_name}' not found. Skipping.")
                continue
            
            # Get both template and output worksheets
            template_worksheet = template_workbook[sheet_name]
            output_worksheet = workbook[sheet_name]
            
            # Use config loader to get unified sheet config
            sheet_config = config_loader.get_sheet_config(sheet_name)
            data_source_indicator = config_loader.get_data_source_type(sheet_name)

            if not data_source_indicator:
                logger.warning(f"Warning: No data source configured for sheet '{sheet_name}'. Skipping.")
                continue

            # --- Processor Factory ---
            processor = None
            if data_source_indicator in ["processed_tables_multi", "processed_tables"]:
                processor = MultiTableProcessor(
                    template_workbook=template_workbook,
                    output_workbook=output_workbook,
                    template_worksheet=template_worksheet,
                    output_worksheet=output_worksheet,
                    sheet_name=sheet_name,
                    sheet_config=sheet_config,
                    config_loader=config_loader,  # Pass config loader instead of raw config dict
                    data_source_indicator=data_source_indicator,
                    invoice_data=invoice_data,
                    cli_args=args,
                    final_grand_total_pallets=final_grand_total_pallets
                )
            else: # Default to single table processor
                processor = SingleTableProcessor(
                    template_workbook=template_workbook,
                    output_workbook=output_workbook,
                    template_worksheet=template_worksheet,
                    output_worksheet=output_worksheet,
                    sheet_name=sheet_name,
                    sheet_config=sheet_config,
                    config_loader=config_loader,  # Pass config loader instead of raw config dict
                    data_source_indicator=data_source_indicator,
                    invoice_data=invoice_data,
                    cli_args=args,
                    final_grand_total_pallets=final_grand_total_pallets
                )
            
            # --- Execute Processing ---
            if processor:
                processing_successful = processor.process()
                if not processing_successful:
                    logger.error(f"--- ERROR occurred while processing sheet '{sheet_name}'. Continuing to next sheet. ---")
                    sheets_failed.append(sheet_name)
                    # Don't break - continue processing other sheets to see all errors
                else:
                    sheets_processed.append(sheet_name)
            else:
                logger.warning(f"Warning: No suitable processor found for sheet '{sheet_name}'. Skipping.")
                sheets_failed.append(sheet_name)

        # --- End of Loop ---
        
        # Summary of processing results
        logger.info(f"=== Processing Summary ===")
        logger.info(f"Sheets processed successfully: {len(sheets_processed)}/{len(sheets_to_process)}")
        if sheets_processed:
            logger.info(f"  Success: {', '.join(sheets_processed)}")
        if sheets_failed:
            logger.error(f"  Failed: {', '.join(sheets_failed)}")
        
        processing_successful = len(sheets_failed) == 0

        if processing_successful:
            logger.info("Saving final workbook...")
            # DEBUG: Check font before saving
            if 'Invoice' in output_workbook.sheetnames:
                ws_debug = output_workbook['Invoice']
                cell_debug = ws_debug['A21']
                logger.debug(f"BEFORE SAVE - A21 font: name={cell_debug.font.name}, size={cell_debug.font.size}, bold={cell_debug.font.bold}")
            output_workbook.save(output_path)
            logger.info(f"Workbook saved successfully: '{output_path}'")
        else:
            logger.error("--- Processing completed with errors. Saving workbook (may be incomplete). ---")
            # DEBUG: Check font before saving
            if 'Invoice' in output_workbook.sheetnames:
                ws_debug = output_workbook['Invoice']
                cell_debug = ws_debug['A21']
                logger.debug(f"BEFORE SAVE - A21 font: name={cell_debug.font.name}, size={cell_debug.font.size}, bold={cell_debug.font.bold}")
            output_workbook.save(output_path)

    except Exception as e:
        logger.error(f"\n--- UNHANDLED ERROR: {e} ---"); traceback.print_exc()
    finally:
        # Close both workbooks
        if template_workbook:
            try: template_workbook.close(); logger.debug("Template workbook closed.")
            except Exception: pass
        if output_workbook:
            try: output_workbook.close(); logger.debug("Output workbook closed.")
            except Exception: pass

    total_time = time.time() - start_time
    logger.info("=== Invoice Generation Finished ===")
    logger.info(f"Total Time: {total_time:.2f} seconds")
    logger.info(f"Completed at: {time.strftime('%H:%M:%S', time.localtime())}")

    # Generate Metadata
    status = "success" if processing_successful else "error"
    error_msg = None if processing_successful else "Processing completed with errors (see logs)"
    
    # Collect replacements log and header info from processor if available
    replacements_log = []
    header_info = {}
    if 'processor' in locals():
        if hasattr(processor, 'replacements_log'):
            replacements_log = processor.replacements_log
        if hasattr(processor, 'header_info'):
            header_info = processor.header_info

    generate_metadata(output_path, status, total_time, sheets_processed, sheets_failed, error_msg, invoice_data, args, replacements_log, header_info)

if __name__ == "__main__":
    # To run this script directly, you might need to adjust Python's path
    # to recognize the 'invoice_generator' package, e.g., by running from the parent directory:
    # python -m invoice_generator.generate_invoice ...
    main()
