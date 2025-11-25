"""
Example: Using Column Mapping with TemplateStateBuilder

This example demonstrates how to use column mapping to handle scenarios where
template columns are filtered/removed during invoice generation (e.g., DAF mode,
custom mode, or conditional column logic).

PROBLEM:
--------
Template has 7 columns: [A, B, C, D, E, F, G]
After filtering (e.g., DAF mode), output only uses: [A, B, D, E, G] (removed C and F)
When restoring template content, without mapping:
  - Template content from column D goes to column D (WRONG - should go to C)
  - Template content from column E goes to column E (WRONG - should go to D)
  - Result: Template content appears in wrong columns

SOLUTION:
---------
Use column mapping to tell TemplateStateBuilder where each template column 
should go in the output.

Mapping: {1: 1, 2: 2, 3: None, 4: 3, 5: 4, 6: None, 7: 5}
  - Column 1 (A) -> Column 1 (A) ✓
  - Column 2 (B) -> Column 2 (B) ✓
  - Column 3 (C) -> None (removed) ✗
  - Column 4 (D) -> Column 3 (C) ✓ SHIFTED
  - Column 5 (E) -> Column 4 (D) ✓ SHIFTED
  - Column 6 (F) -> None (removed) ✗
  - Column 7 (G) -> Column 5 (E) ✓ SHIFTED
"""

from openpyxl import load_workbook
from invoice_generator.builders.template_state_builder import TemplateStateBuilder
import logging

logger = logging.getLogger(__name__)

def example_with_column_mapping():
    """
    Example: Setting up column mapping for DAF mode invoice generation
    """
    
    # 1. Load template and output workbooks
    template_wb = load_workbook("template/JF.xlsx")
    output_wb = load_workbook("template/JF.xlsx")  # Copy for writing
    
    template_ws = template_wb['Invoice']
    output_ws = output_wb['Invoice']
    
    # 2. Create TemplateStateBuilder with template worksheet
    template_state = TemplateStateBuilder(
        worksheet=template_ws,
        num_header_cols=7,  # Template has 7 columns
        header_end_row=21,  # Header ends at row 21
        footer_start_row=50,  # Footer starts at row 50
        debug=True  # Enable debug logging
    )
    
    # 3. Define column mapping based on which columns are kept
    # In this example, columns 3 and 6 are removed (None), others are shifted
    column_mapping = {
        1: 1,      # col_static: A -> A (kept)
        2: 2,      # col_po: B -> B (kept)
        3: None,   # col_item: C -> removed
        4: 3,      # col_desc: D -> C (shifted left)
        5: 4,      # col_qty_sf: E -> D (shifted left)
        6: None,   # col_unit_price: F -> removed
        7: 5       # col_amount: G -> E (shifted left)
    }
    
    # 4. Apply the column mapping
    template_state.set_column_mapping(column_mapping)
    logger.info("Column mapping applied: template will shift content to match output columns")
    
    # 5. Build the data table (this creates the actual invoice data)
    # ... (data table building code here) ...
    
    # 6. Restore header with column mapping
    template_state.restore_header_only(
        target_worksheet=output_ws,
        actual_num_cols=5  # Output has 5 columns (not 7)
    )
    logger.info("Header restored with column shifts applied")
    
    # 7. Restore footer with column mapping
    template_state.restore_footer_only(
        target_worksheet=output_ws,
        footer_start_row=60,  # Footer starts after data
        actual_num_cols=5  # Output has 5 columns
    )
    logger.info("Footer restored with column shifts applied")
    
    # 8. Save output
    output_wb.save("result_with_column_mapping.xlsx")
    logger.info("Invoice saved with correctly shifted template content")
    
    # Clean up
    template_wb.close()
    output_wb.close()


def generate_column_mapping_from_filtered_columns(
    template_columns: list,
    filtered_columns: list
) -> dict:
    """
    Helper function to automatically generate column mapping.
    
    Args:
        template_columns: List of all column IDs from template
                         Example: ['col_static', 'col_po', 'col_item', ...]
        filtered_columns: List of column IDs after filtering
                         Example: ['col_static', 'col_po', 'col_desc', ...]
    
    Returns:
        Column mapping dict {template_col_index: output_col_index or None}
    """
    mapping = {}
    output_index = 1
    
    for template_index, col_id in enumerate(template_columns, start=1):
        if col_id in filtered_columns:
            # Column is kept - map to current output position
            mapping[template_index] = output_index
            output_index += 1
        else:
            # Column is removed - map to None
            mapping[template_index] = None
    
    return mapping


def example_auto_generate_mapping():
    """
    Example: Automatically generating column mapping from filtered columns
    """
    
    # Original template columns
    template_columns = [
        'col_static', 'col_po', 'col_item', 'col_desc', 
        'col_qty_sf', 'col_unit_price', 'col_amount'
    ]
    
    # After DAF filtering (removed col_item and col_unit_price)
    filtered_columns = [
        'col_static', 'col_po', 'col_desc', 
        'col_qty_sf', 'col_amount'
    ]
    
    # Generate mapping automatically
    mapping = generate_column_mapping_from_filtered_columns(
        template_columns=template_columns,
        filtered_columns=filtered_columns
    )
    
    print("Auto-generated mapping:", mapping)
    # Output: {1: 1, 2: 2, 3: None, 4: 3, 5: 4, 6: None, 7: 5}
    
    # Use this mapping with TemplateStateBuilder
    # template_state.set_column_mapping(mapping)


if __name__ == "__main__":
    logging.basicConfig(level=logging.DEBUG)
    
    print("=" * 70)
    print("Example 1: Manual column mapping")
    print("=" * 70)
    # example_with_column_mapping()  # Uncomment to run
    
    print("\n" + "=" * 70)
    print("Example 2: Auto-generate mapping from filtered columns")
    print("=" * 70)
    example_auto_generate_mapping()
