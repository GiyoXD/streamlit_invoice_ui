"""
Example: Using BuilderConfigResolver with LayoutBuilder

This example demonstrates the modern approach to using LayoutBuilder
with the BuilderConfigResolver pattern, eliminating legacy data source logic.
"""

from pathlib import Path
from openpyxl import load_workbook

from invoice_generator.config.config_loader import BundledConfigLoader
from invoice_generator.config.builder_config_resolver import BuilderConfigResolver
from invoice_generator.builders.layout_builder import LayoutBuilder
import logging
logger = logging.getLogger(__name__)


def build_invoice_with_resolver_pattern(config_path: Path, template_path: Path, invoice_data: dict, args):
    """
    Example of the MODERN approach using BuilderConfigResolver.
    
    This eliminates all legacy data source resolution logic from LayoutBuilder
    by using the resolver to prepare all configuration bundles.
    """
    # 1. Load bundled configuration
    config_loader = BundledConfigLoader(config_path)
    
    # 2. Load workbook and template
    template_workbook = load_workbook(template_path)
    output_workbook = load_workbook(template_path)  # Or create new
    
    sheet_name = 'Invoice'
    template_worksheet = template_workbook[sheet_name]
    output_worksheet = output_workbook[sheet_name]
    
    # 3. Create resolver instance
    # The resolver encapsulates ALL config extraction logic
    resolver = BuilderConfigResolver(
        config_loader=config_loader,
        sheet_name=sheet_name,
        worksheet=output_worksheet,
        args=args,
        invoice_data=invoice_data,
        pallets=31
    )
    
    # 4. Get bundles using the new convenience method
    # This method merges layout_config and data_config for LayoutBuilder
    style_config, context_config, layout_config = resolver.get_layout_bundles_with_data()
    
    # 5. Create LayoutBuilder with resolved bundles
    # NO legacy parameters needed - everything comes from bundles
    layout_builder = LayoutBuilder(
        workbook=output_workbook,
        worksheet=output_worksheet,
        template_worksheet=template_worksheet,
        style_config=style_config,
        context_config=context_config,
        layout_config=layout_config
    )
    
    # 6. Build the layout
    success = layout_builder.build()
    
    if success:
        logger.info(f"✅ Layout built successfully using resolver pattern")
        logger.info(f"   Data source type: {layout_config.get('data_source_type')}")
        logger.info(f"   Next row after footer: {layout_builder.next_row_after_footer}")
    else:
        logger.critical(f"❌ Layout build failed")
    
    return success


def build_invoice_legacy_approach(template_path: Path, sheet_config: dict, invoice_data: dict, args):
    """
    Example of the LEGACY approach (DEPRECATED).
    
    This approach requires passing individual parameters and relies on
    LayoutBuilder to extract data sources using legacy logic.
    
    warning!! NOT RECOMMENDED - Use resolver pattern above instead
    """
    # Load workbook
    template_workbook = load_workbook(template_path)
    output_workbook = load_workbook(template_path)
    
    sheet_name = 'Invoice'
    template_worksheet = template_workbook[sheet_name]
    output_worksheet = output_workbook[sheet_name]
    
    # Create LayoutBuilder with individual parameters
    # This triggers legacy data source resolution logic
    layout_builder = LayoutBuilder(
        workbook=output_workbook,
        worksheet=output_worksheet,
        template_worksheet=template_worksheet,
        sheet_name=sheet_name,
        sheet_config=sheet_config,  # Individual config dict
        invoice_data=invoice_data,  # Raw invoice data
        args=args,
        final_grand_total_pallets=31
    )
    
    # Build the layout (will use legacy path)
    success = layout_builder.build()

    logger.warning(f"warning!! Used legacy approach - consider migrating to resolver pattern")

    return success


def build_multi_table_with_resolver(config_path: Path, template_path: Path, invoice_data: dict, args):
    """
    Example of multi-table processing using BuilderConfigResolver.
    
    Demonstrates how to use table_key parameter for multi-table sheets
    like "Packing list" which processes multiple table chunks.
    """
    config_loader = BundledConfigLoader(config_path)
    template_workbook = load_workbook(template_path)
    output_workbook = load_workbook(template_path)
    
    sheet_name = 'Packing list'
    template_worksheet = template_workbook[sheet_name]
    output_worksheet = output_workbook[sheet_name]
    
    # Process table '1'
    resolver_table1 = BuilderConfigResolver(
        config_loader=config_loader,
        sheet_name=sheet_name,
        worksheet=output_worksheet,
        args=args,
        invoice_data=invoice_data,
        pallets=20  # Pallets for table 1
    )
    
    # Get bundles for first table
    style_config, context_config, layout_config = resolver_table1.get_layout_bundles_with_data(table_key='1')
    
    layout_builder_table1 = LayoutBuilder(
        workbook=output_workbook,
        worksheet=output_worksheet,
        template_worksheet=template_worksheet,
        style_config=style_config,
        context_config=context_config,
        layout_config=layout_config
    )
    
    success = layout_builder_table1.build()
    current_row = layout_builder_table1.next_row_after_footer
    
    print(f"✅ Table '1' complete. Next row: {current_row}, Pallets: 20")
    
    # Process table '2' at the next available row
    resolver_table2 = BuilderConfigResolver(
        config_loader=config_loader,
        sheet_name=sheet_name,
        worksheet=output_worksheet,
        args=args,
        invoice_data=invoice_data,
        pallets=11  # Pallets for table 2
    )
    
    style_config, context_config, layout_config = resolver_table2.get_layout_bundles_with_data(table_key='2')
    
    # Update header_row to write second table after first
    layout_config['sheet_config']['header_row'] = current_row + 1
    
    layout_builder_table2 = LayoutBuilder(
        workbook=output_workbook,
        worksheet=output_worksheet,
        template_worksheet=template_worksheet,
        style_config=style_config,
        context_config=context_config,
        layout_config=layout_config,
        skip_template_header_restoration=True,  # Only restore once
        skip_template_footer_restoration=False  # Restore after last table
    )
    
    success = layout_builder_table2.build()
    
    print(f"✅ Table '2' complete. Next row: {layout_builder_table2.next_row_after_footer}, Pallets: 11")
    
    return success


if __name__ == '__main__':
    """
    Quick reference for LayoutBuilder + Resolver integration
    """
    print("=" * 70)
    print("LayoutBuilder + BuilderConfigResolver Pattern")
    print("=" * 70)
    print()
    print("MODERN APPROACH (Recommended):")
    print("  1. Load config with BundledConfigLoader")
    print("  2. Create BuilderConfigResolver instance")
    print("  3. Call resolver.get_layout_bundles_with_data()")
    print("  4. Pass bundles to LayoutBuilder via style_config, context_config, layout_config")
    print("  5. Call layout_builder.build()")
    print()
    print("BENEFITS:")
    print("  [OK]OK]OK] Centralized data source resolution")
    print("  [OK]OK]OK] No duplicate extraction logic")
    print("  [OK]OK]OK] Cleaner separation of concerns")
    print("  [OK]OK]OK] Easier to test and maintain")
    print("  [OK]OK]OK] Supports bundled config v2.1+")
    print()
    print("LEGACY APPROACH (Deprecated):")
    print("  warning!! Direct parameter passing with individual dicts")
    print("  warning!! Triggers legacy data source resolution in LayoutBuilder")
    print("  warning!! Harder to maintain and test")
    print()
    print("=" * 70)
