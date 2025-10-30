"""
Header detection functionality for Excel analysis tool.

This module provides the HeaderDetector class that searches for specific header keywords
and determines the start row for data insertion.
"""

from typing import List, Optional, Dict
import json
import os
from pathlib import Path
from openpyxl.worksheet.worksheet import Worksheet
from models.data_models import HeaderMatch


class HeaderDetector:
    """Detects header keywords and calculates start row positions."""
    
    def __init__(self, quantity_mode: bool = False, mapping_config: Optional[Dict] = None):
        """Initialize the HeaderDetector.
        
        Args:
            quantity_mode: If True, adds PCS and SQFT columns for packing list sheets
            mapping_config: Optional mapping configuration dictionary
        """
        self.quantity_mode = quantity_mode
        self.mapping_config = mapping_config
        self.header_keywords = self._load_header_keywords()
        # Load exact headers from mapping config if available
        if self.mapping_config:
            self.exact_headers = list(self.mapping_config.get('header_text_mappings', {}).get('mappings', {}).keys())
        else:
            self.exact_headers = []
    
    def _load_header_keywords(self) -> List[str]:
        """Load header keywords from mapping config or use defaults."""
        keywords = set()
        
        # Try to load from mapping config first
        if self.mapping_config:
            try:
                header_mappings = self.mapping_config.get('header_text_mappings', {}).get('mappings', {})
                for header in header_mappings.keys():
                    # Extract base keywords from headers
                    keywords.update(self._extract_keywords_from_header(header))
            except Exception as e:
                import logging
                logging.warning(f"Could not load keywords from mapping config: {e}")
                print(f"Warning: Could not load keywords from mapping config: {e}")
        
        # If no keywords loaded from config, use defaults
        if not keywords:
            keywords.update([
                "P.O", "ITEM", "Description", "Quantity", "Amount",
                "Mark", "Unit price", "Price", "Total", "Weight", "CBM", "Pallet",
                "Remarks", "HS CODE", "Name", "Commodity", "Goods", "Product",
                "PCS", "SF", "No.", "N.W", "G.W", "Net", "Gross", "FCA"
            ])
        
        return list(keywords)
    
    def _extract_keywords_from_header(self, header: str) -> List[str]:
        """Extract meaningful keywords from a header string."""
        import re
        
        keywords = []
        header_lower = header.lower()
        
        # Split by common separators and extract meaningful words
        filtered_words = [word for word in words if word not in generic_words]
        keywords.extend(filtered_words)
        
        # Add common combinations that might appear in headers
        if 'unit' in header_lower and 'price' in header_lower:
            keywords.append('unit price')
        
        # Add common combinations that might appear in headers
        if 'unit' in header_lower and 'price' in header_lower:
            keywords.append('unit price')
        if 'gross' in header_lower and 'weight' in header_lower:
            keywords.append('gross weight')
        if 'net' in header_lower and 'weight' in header_lower:
            keywords.append('net weight')
        if 'p.o' in header_lower:
            keywords.append('p.o')
        if 'hs' in header_lower and 'code' in header_lower:
            keywords.append('hs code')
        
        return keywords
    
    def find_headers(self, worksheet: Worksheet) -> List[HeaderMatch]:
        """
        IMPROVED: Search for header row using BOLD formatting detection.
        Only processes these sheet types (case insensitive):
        - Invoice
        - Packing list
        - Detail packing list
        - Contract
        
        The row with the most bold cells is identified as the header row.
        If "Quantity" header is detected, automatically handles 2-row header structure.
        
        Args:
            worksheet: The openpyxl worksheet to analyze
            
        Returns:
            List of HeaderMatch objects containing keyword, row, and column positions
        """
        # VALIDATE: Only process specific sheet types
        sheet_name = worksheet.title.lower().strip()
        valid_sheets = [
            'invoice',
            'packing list',
            'detail packing list',
            'contract'
        ]
        
        is_valid_sheet = any(valid_name in sheet_name for valid_name in valid_sheets)
        
        if not is_valid_sheet:
            print(f"[HEADER_DETECTION] Sheet '{worksheet.title}' is not a valid type")
            print(f"[HEADER_DETECTION] Valid types: Invoice, Packing list, Detail packing list, Contract")
            return []
        
        print(f"[HEADER_DETECTION] Processing sheet: {worksheet.title}")
        
        max_rows_to_check = 30
        
        # Find header row using bold detection
        header_row_found = self._find_header_row_by_bold(worksheet, max_rows_to_check)
        
        if not header_row_found:
            print("[HEADER_DETECTION] No bold header row found, using fallback keyword detection")
            header_row_found = self._find_header_row_by_keywords(worksheet, max_rows_to_check)
        
        if not header_row_found:
            print("[HEADER_DETECTION] No header row found!")
            return []
        
        print(f"[HEADER_DETECTION] Selected header row: {header_row_found}")
        
        # Check if this is a 2-row header by looking for "Quantity" keyword
        is_double_header = self._is_double_header_quantity_based(worksheet, header_row_found)
        
        if is_double_header:
            print(f"[HEADER_DETECTION] Detected 2-row header structure (Quantity found)")
            header_matches = self._extract_double_header(worksheet, header_row_found)
        else:
            print(f"[HEADER_DETECTION] Detected single-row header structure")
            header_matches = self._extract_all_headers_from_row(worksheet, header_row_found)
        
        # Apply quantity mode enhancement if enabled
        if self.quantity_mode:
            header_matches = self._apply_quantity_mode_enhancement(header_matches, worksheet)
        
        return header_matches
    
    def _find_header_row_by_bold(self, worksheet: Worksheet, max_rows: int) -> Optional[int]:
        """
        Find the header row by detecting the row with the MOST bold cells.
        SIMPLE RULE: Row with most bold cells = header row.
        Also validates that the row contains HEADER TEXT (not data values).
        
        Args:
            worksheet: The worksheet to analyze
            max_rows: Maximum number of rows to check
            
        Returns:
            Row number of the header row, or None if not found
        """
        candidates = []
        
        for row_num in range(1, min(max_rows + 1, worksheet.max_row + 1)):
            bold_count = 0
            filled_count = 0
            text_count = 0
            numeric_count = 0
            
            # Count bold cells and analyze cell content
            for col in range(1, min(21, worksheet.max_column + 1)):
                cell = worksheet.cell(row=row_num, column=col)
                
                if cell.value is not None:
                    cell_value = str(cell.value).strip()
                    if cell_value:
                        filled_count += 1
                        
                        # Check if cell is bold
                        if cell.font and cell.font.bold:
                            bold_count += 1
                        
                        # Check if cell contains text or numbers
                        # Remove common punctuation and check
                        clean_value = cell_value.replace(',', '').replace('$', '').replace('.', '')
                        if clean_value.replace('-', '').isdigit():
                            numeric_count += 1
                        else:
                            text_count += 1
            
            # Row must have:
            # 1. At least 3 bold cells
            # 2. More text than numbers (headers are text labels, not data)
            if bold_count >= 3 and text_count > numeric_count:
                candidates.append({
                    'row': row_num,
                    'bold_count': bold_count,
                    'filled_count': filled_count,
                    'text_count': text_count,
                    'numeric_count': numeric_count
                })
        
        if not candidates:
            return None
        
        # Sort by bold count ONLY - highest wins
        candidates.sort(key=lambda x: x['bold_count'], reverse=True)
        
        # Debug output
        print(f"[BOLD_DETECTION] Found {len(candidates)} candidates with bold cells:")
        for i, c in enumerate(candidates[:5]):  # Show top 5
            marker = " ← SELECTED (MOST BOLD)" if i == 0 else ""
            print(f"  Row {c['row']:2d}: {c['bold_count']:2d} bold cells, "
                  f"{c['text_count']} text, {c['numeric_count']} numeric{marker}")
        
        return candidates[0]['row']
    
    def _find_header_row_by_keywords(self, worksheet: Worksheet, max_rows: int) -> Optional[int]:
        """
        Fallback method: Find header row by keyword matching.
        Used when bold detection fails.
        
        Args:
            worksheet: The worksheet to analyze
            max_rows: Maximum number of rows to check
            
        Returns:
            Row number of the header row, or None if not found
        """
        for row_num in range(1, min(max_rows + 1, worksheet.max_row + 1)):
            keyword_count = 0
            
            for col in range(1, min(21, worksheet.max_column + 1)):
                cell = worksheet.cell(row=row_num, column=col)
                
                if cell.value is not None:
                    cell_value = str(cell.value).strip()
                    
                    # Check if cell contains any header keyword
                    for keyword in self.header_keywords:
                        if self._matches_keyword(cell_value, keyword):
                            keyword_count += 1
                            break
            
            # If we found at least 3 keywords in this row, it's likely the header
            if keyword_count >= 3:
                print(f"[KEYWORD_DETECTION] Found header row {row_num} with {keyword_count} keywords")
                return row_num
        
        return None
    
    def _is_double_header_quantity_based(self, worksheet: Worksheet, header_row: int) -> bool:
        """
        STRICT: Determine if this is a 2-row header with multiple validation checks.
        Only returns True if ALL conditions are met:
        1. "Quantity" keyword exists in header row
        2. Next row has SUB-HEADERS (short text like PCS, SF), NOT data
        3. Next row has bold cells (formatted like headers)
        4. Next row cells are SHORT (2-5 chars like "PCS", "SF")
        
        Args:
            worksheet: The worksheet to analyze
            header_row: The detected header row number
            
        Returns:
            True if this is a 2-row header (all conditions met), False otherwise
        """
        # CHECK 1: "Quantity" keyword must exist
        has_quantity = False
        for col in range(1, min(21, worksheet.max_column + 1)):
            cell = worksheet.cell(row=header_row, column=col)
            
            if cell.value is not None:
                cell_value = str(cell.value).strip().lower()
                
                # Check for "Quantity" or "Qty" variations
                if 'quantity' in cell_value or cell_value == 'qty' or 'qty' in cell_value:
                    has_quantity = True
                    print(f"[DOUBLE_HEADER_DETECTION] ✓ Found 'Quantity' at row {header_row}, col {col}")
                    break
        
        if not has_quantity:
            print(f"[DOUBLE_HEADER_DETECTION] ✗ No 'Quantity' keyword found → Single-row header")
            return False
        
        # CHECK 2: Next row must exist
        if header_row >= worksheet.max_row:
            print(f"[DOUBLE_HEADER_DETECTION] ✗ No next row → Single-row header")
            return False
        
        next_row = header_row + 1
        
        # CHECK 3: Analyze next row content
        text_cells = 0
        numeric_cells = 0
        bold_cells = 0
        short_text_cells = 0  # NEW: Count short text (like PCS, SF)
        long_text_cells = 0   # NEW: Count long text (like descriptions)
        
        for col in range(1, min(21, worksheet.max_column + 1)):
            cell = worksheet.cell(row=next_row, column=col)
            if cell.value is not None:
                cell_value = str(cell.value).strip()
                if cell_value:
                    # Check if bold
                    if cell.font and cell.font.bold:
                        bold_cells += 1
                    
                    # Check if numeric or text
                    clean_value = cell_value.replace(',', '').replace('$', '').replace('.', '')
                    if clean_value.replace('-', '').isdigit():
                        numeric_cells += 1
                    else:
                        text_cells += 1
                        # Check length - sub-headers are SHORT (PCS, SF, KG)
                        if len(cell_value) <= 5:
                            short_text_cells += 1
                        else:
                            long_text_cells += 1
        
        # CHECK 4: Next row must have more text than numbers
        if text_cells == 0 or numeric_cells >= text_cells:
            print(f"[DOUBLE_HEADER_DETECTION] ✗ Next row has {text_cells} text, {numeric_cells} numeric → Looks like data row")
            return False
        
        # CHECK 5: Next row must have at least some bold cells (formatted like headers)
        if bold_cells < 2:
            print(f"[DOUBLE_HEADER_DETECTION] ✗ Next row has only {bold_cells} bold cells → Not a header row")
            return False
        
        # CHECK 6: Next row should have SHORT text (sub-headers), not LONG text (descriptions)
        if long_text_cells > short_text_cells:
            print(f"[DOUBLE_HEADER_DETECTION] ✗ Next row has {long_text_cells} long text cells → Looks like data descriptions")
            return False
        
        # ALL CHECKS PASSED - This is a 2-row header!
        print(f"[DOUBLE_HEADER_DETECTION] ✓ Confirmed 2-row header:")
        print(f"  Next row: {text_cells} text ({short_text_cells} short, {long_text_cells} long), "
              f"{numeric_cells} numeric, {bold_cells} bold cells")
        return True
    
    def calculate_start_row(self, header_positions: List[HeaderMatch]) -> int:
        """
        Calculate the start row where headers begin.
        
        Args:
            header_positions: List of HeaderMatch objects
            
        Returns:
            The row number where headers start (min_header_row)
        """
        if not header_positions:
            return 1  # Default to row 1 if no headers found
        
        # Find the minimum header row (where headers start)
        min_header_row = min(match.row for match in header_positions)
        return min_header_row
    
    def _extract_all_headers_from_row(self, worksheet: Worksheet, header_row: int) -> List[HeaderMatch]:
        """
        Extract all non-empty headers from the specified row.
        
        Args:
            worksheet: The openpyxl worksheet to analyze
            header_row: The row number containing headers
            
        Returns:
            List of HeaderMatch objects for all headers in the row
        """
        header_matches = []
        
        # Get the specific row and extract all non-empty cells
        for cell in worksheet[header_row]:
            if cell.value is not None:
                cell_value = str(cell.value).strip()
                if cell_value:  # Only include non-empty values
                    header_match = HeaderMatch(
                        keyword=cell_value,  # Use the actual cell value as the keyword
                        row=cell.row,
                        column=cell.column
                    )
                    header_matches.append(header_match)
        
        return header_matches
    
    def _apply_quantity_mode_enhancement(self, header_matches: List[HeaderMatch], worksheet: Worksheet) -> List[HeaderMatch]:
        """
        Apply quantity mode enhancement for packing list sheets.
        Adds PCS and SQFT columns after Quantity column.
        
        Args:
            header_matches: Original list of header matches
            worksheet: The worksheet being analyzed
            
        Returns:
            Enhanced list of header matches with PCS and SQFT columns
        """
        # Check if this is a packing list sheet
        sheet_name = worksheet.title.lower()
        if not any(keyword in sheet_name for keyword in ['packing', 'pkl', 'packing list']):
            return header_matches  # Not a packing list, return original
        
        # Find the Quantity column
        quantity_match = None
        for match in header_matches:
            if 'quantity' in match.keyword.lower():
                quantity_match = match
                break
        
        if not quantity_match:
            return header_matches  # No quantity column found
        
        # Create enhanced header list with original headers
        enhanced_headers = header_matches.copy()
        
        # Add PCS and SQFT in the row BELOW the Quantity header
        # PCS: same column as Quantity, but row + 1
        pcs_header = HeaderMatch(
            keyword="PCS",
            row=quantity_match.row + 1,
            column=quantity_match.column
        )
        enhanced_headers.append(pcs_header)
        
        # SQFT: same row as PCS, but next column
        sqft_header = HeaderMatch(
            keyword="SF", 
            row=quantity_match.row + 1,
            column=quantity_match.column + 1
        )
        enhanced_headers.append(sqft_header)
        
        return enhanced_headers
    
    def _extract_double_header(self, worksheet: Worksheet, header_row: int) -> List[HeaderMatch]:
        """
        Extract headers from a two-row header structure.
        
        Args:
            worksheet: The openpyxl worksheet to analyze
            header_row: The first row of the header
            
        Returns:
            List of HeaderMatch objects for all headers in both rows
        """
        header_matches = []
        
        # Extract headers from the first row
        for cell in worksheet[header_row]:
            if cell.value is not None:
                cell_value = str(cell.value).strip()
                if cell_value:
                    header_match = HeaderMatch(
                        keyword=cell_value,
                        row=cell.row,
                        column=cell.column
                    )
                    header_matches.append(header_match)
        
        # Extract headers from the second row (header_row + 1)
        second_row = header_row + 1
        for cell in worksheet[second_row]:
            if cell.value is not None:
                cell_value = str(cell.value).strip()
                if cell_value:
                    header_match = HeaderMatch(
                        keyword=cell_value,
                        row=cell.row,
                        column=cell.column
                    )
                    header_matches.append(header_match)
        
        return header_matches
    
    def _matches_keyword(self, cell_value: str, keyword: str) -> bool:
        """
        Check if a cell value matches a header keyword.
        Uses strict matching to avoid false positives - cell should be primarily the keyword.
        
        Args:
            cell_value: The cell value to check
            keyword: The keyword to match against
            
        Returns:
            True if the cell value is primarily the keyword (case-insensitive)
        """
        cell_lower = cell_value.lower().strip()
        keyword_lower = keyword.lower()
        
        # Exact match
        if cell_lower == keyword_lower:
            return True
            
        # Allow some common variations but keep it strict
        # Remove common punctuation and extra spaces
        import re
        cell_clean = re.sub(r'[^\w\s]', ' ', cell_lower)
        cell_clean = ' '.join(cell_clean.split())  # normalize whitespace
        
        keyword_clean = re.sub(r'[^\w\s]', ' ', keyword_lower)
        keyword_clean = ' '.join(keyword_clean.split())
        
        # Check if cleaned versions match
        if cell_clean == keyword_clean:
            return True
            
        # For very short cells (likely headers), allow if keyword is majority of content
        if len(cell_lower) <= 20 and keyword_lower in cell_lower:
            # Calculate similarity - keyword should be significant portion
            similarity = len(keyword_lower) / len(cell_lower)
            return similarity >= 0.6
        
        return False